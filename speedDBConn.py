import os
import datetime
import speedTest as spd
import xlwings as xw
from openpyxl import *


path = os.path.basename(r"C:\Users\xacc\Desktop\speedTestsDB.xlsx")
# wb = load_workbook(path)

# Creates list for day and time. 0-4 = year, 6-7 = month, 9-10 = day, 12-13 = hour, 15-16 = minutes, 18-19 = seconds, 21-26 = ms
dateTimeStr = str(datetime.datetime.now())
hoursMinutes = dateTimeStr[11:16]
date = dateTimeStr[0:10]

if int(hoursMinutes[:2]) > 12:
    hoursMinutes = str(int(hoursMinutes[:2]) - 12) + hoursMinutes[2:] + "pm"
else:
    hoursMinutes = str(int(hoursMinutes[:2]) - 12) + hoursMinutes[2:] + "am"

finalDateTime = "{0}/{1}/{2} -- {3}".format(date[5:7], date[8:10], date[0:4], hoursMinutes)

print(finalDateTime)

def write_data_to_excel(workbooklocation,sheetname,columnletter):
    wb2 = xw.Book(workbooklocation)
    x = wb2.sheets[sheetname].range(columnletter + str(wb2.sheets[sheetname].cells.last_cell.row)).end('up').row + 1
    cell = columnletter + str(x)
    print(cell)
    return cell

wb2 = xw.Book('speedTestsDB.xlsx')
ws = wb2.sheets['speedTests']
last_row = write_data_to_excel('speedTestsDB.xlsx', 'speedTests', 'B')
print(last_row[1:])

# ws = wb['speedTests']
# Date/Time
wc1 = ws.cell(int(last_row[1:]), 2)
wc1.value = finalDateTime
# # Ping
wc2 = ws.cell(int(last_row[1:]),3)
wc2.value = spd.s.results.ping
# # Download
wc3 = ws.cell(int(last_row[1:]),4)
wc3.value = spd.mbpsDown
# # Upload
wc4 = ws.cell(int(last_row[1:]),5)
wc4.value = spd.mbpsUp

wb.save(path)